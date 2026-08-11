# -*- coding: utf-8 -*-
"""Build meeting task Excel report from JSON.

Headers based on corporate status-meeting template + column «Примечание».

Run from repo root:
  python tools/build_meeting_task_report_xlsx.py \\
    --json docs/04-registry/meetings/example.tasks.json \\
    --out docs/04-registry/meetings/example-tasks.xlsx

Optional:
  --template platform/templates/meeting-tasks.xlsx  (copy header styling from template)
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = [
    "№",
    "Задача",
    "Ответственный",
    "Срок",
    "Основание (тайм-код)",
    "Выполнение.",
    "Примечание",
]

JSON_KEYS = [
    "task_num",
    "task",
    "responsible",
    "deadline",
    "timecode_basis",
    "completion_status",
    "notes",
]

HEADER_FILL = PatternFill("solid", fgColor="D9C4A0")
THIN = Side(style="thin", color="000000")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

COL_WIDTHS = [6, 52, 22, 28, 22, 36, 28]


def row_from_obj(obj: dict[str, Any]) -> list[Any]:
    return [obj.get(k, "") for k in JSON_KEYS]


def style_sheet(ws) -> None:
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for col_idx, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
        cell.alignment = CENTER
        cell.border = CELL_BORDER

    ws.freeze_panes = "A2"


def build_xlsx(rows: list[dict[str, Any]], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Задачи"

    style_sheet(ws)

    for r_idx, obj in enumerate(rows, start=2):
        values = row_from_obj(obj)
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = CELL_BORDER
            if c_idx in (2, 6, 7):
                cell.alignment = WRAP
            elif c_idx == 1:
                cell.alignment = CENTER
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def write_header_template(out_path: Path) -> None:
    """Write header-only xlsx for platform/templates/meeting-tasks.xlsx."""
    build_xlsx([], out_path)


def load_rows(json_path: Path) -> list[dict[str, Any]]:
    data = json.loads(json_path.read_text(encoding="utf-8"))
    if isinstance(data, dict) and "rows" in data:
        return data["rows"]
    if isinstance(data, list):
        return data
    raise ValueError("JSON must be a list of row objects or {rows: [...]}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build meeting task report xlsx")
    parser.add_argument("--json", type=Path, help="Input JSON file")
    parser.add_argument("--out", type=Path, required=True, help="Output xlsx path")
    parser.add_argument(
        "--header-only",
        action="store_true",
        help="Write header row only (for platform template)",
    )
    args = parser.parse_args()

    if args.header_only:
        write_header_template(args.out)
        print(f"Wrote header template -> {args.out}")
        return

    if not args.json:
        parser.error("--json is required unless --header-only")

    rows = load_rows(args.json)
    build_xlsx(rows, args.out)
    print(f"Wrote {len(rows)} rows -> {args.out}")


if __name__ == "__main__":
    main()
